"""
Excel 导出服务
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from models.visitor import Visitor, Companion
from models.admin import ApprovalRecord
from services.crypto import aes_decrypt, mask_id_number, mask_phone
from config import get_config

config = get_config()


def _set_cell_style(cell, bold=False, align_center=True, is_header=False):
    """设置单元格样式"""
    if is_header:
        cell.font = Font(bold=True, size=11, name='微软雅黑')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, size=11, name='微软雅黑', color='FFFFFF')
    else:
        cell.font = Font(size=10, name='微软雅黑')

    if align_center:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    cell.border = thin_border


def export_visitors_to_excel():
    """
    导出访客登记记录为Excel
    返回 (bytes, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '访客登记记录'

    # 表头
    headers = [
        '序号', '姓名', '手机号', '身份证号', '接待人', '接待人部门',
        '访问开始时间', '访问结束时间', '访问地点', '访问目的',
        '来访人数', '是否携带设备', '设备名称型号', '车牌号',
        '同行人信息', '审批状态', '登记时间'
    ]

    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _set_cell_style(cell, is_header=True)

    # 查询所有访客记录（包括已过期的）
    visitors = Visitor.query.order_by(Visitor.created_at.desc()).all()

    for idx, visitor in enumerate(visitors, 1):
        row = idx + 1
        decrypted_phone = aes_decrypt(visitor.phone) if visitor.phone else ''
        decrypted_id = aes_decrypt(visitor.id_number) if visitor.id_number else ''

        # 脱敏处理
        masked_phone = mask_phone(decrypted_phone)
        masked_id = mask_id_number(decrypted_id)

        # 同行人信息
        companion_info = ''
        if visitor.companions:
            companions = []
            for c in visitor.companions:
                c_id = aes_decrypt(c.id_number)
                companions.append(f'{c.name}({mask_id_number(c_id)})')
            companion_info = '; '.join(companions)

        # 状态中文
        status_map = {
            'pending': '待审批',
            'level1_approved': '一级已通过',
            'approved': '已通过',
            'rejected': '已拒绝',
        }

        data = [
            idx,
            visitor.name,
            masked_phone,
            masked_id,
            visitor.host_name,
            visitor.host_department,
            visitor.visit_start.strftime('%Y-%m-%d %H:%M') if visitor.visit_start else '',
            visitor.visit_end.strftime('%Y-%m-%d %H:%M') if visitor.visit_end else '',
            visitor.visit_location,
            visitor.visit_purpose,
            visitor.visitor_count,
            '是' if visitor.has_device else '否',
            visitor.device_info or '',
            '; '.join(visitor.get_license_plates()) or '',
            companion_info,
            status_map.get(visitor.status, visitor.status),
            visitor.created_at.strftime('%Y-%m-%d %H:%M:%S') if visitor.created_at else '',
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            _set_cell_style(cell)

    # 调整列宽
    col_widths = [6, 10, 14, 22, 10, 12, 18, 18, 15, 25, 8, 10, 20, 12, 30, 12, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'访客登记记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return output.getvalue(), filename


def export_approvals_to_excel():
    """
    导出审批记录为Excel
    返回 (bytes, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '审批记录'

    headers = [
        '序号', '访客姓名', '接待人部门', '审批人', '审批人角色',
        '审批结果', '审批意见', '审批时间'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _set_cell_style(cell, is_header=True)

    records = db.session.query(
        ApprovalRecord, Visitor
    ).join(
        Visitor, ApprovalRecord.visitor_id == Visitor.id
    ).order_by(ApprovalRecord.approved_at.desc()).all()

    for idx, (record, visitor) in enumerate(records, 1):
        row = idx + 1
        role_map = {'level1': '一级审批人', 'level2': '二级审批人'}
        result_map = {'approved': '通过', 'rejected': '拒绝'}

        data = [
            idx,
            visitor.name,
            visitor.host_department,
            record.approver_name,
            role_map.get(record.approver_role, record.approver_role),
            result_map.get(record.result, record.result),
            record.comment or '',
            record.approved_at.strftime('%Y-%m-%d %H:%M:%S') if record.approved_at else '',
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            _set_cell_style(cell)

    col_widths = [6, 10, 12, 10, 12, 10, 30, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'审批记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return output.getvalue(), filename


def encrypt_excel(data: bytes, password: str) -> bytes:
    """
    为Excel文件设置打开密码
    使用 msoffcrypto 实现真正的文件级加密（password to open）
    注意：openpyxl 的 workbookPassword 只是工作表结构保护密码，不是文件打开密码
    """
    import io as io_module
    import msoffcrypto

    temp_input = io_module.BytesIO(data)
    office_file = msoffcrypto.OfficeFile(temp_input)

    temp_output = io_module.BytesIO()
    office_file.encrypt(password, temp_output)
    temp_output.seek(0)

    return temp_output.getvalue()


# Import for the query join
from models import db
